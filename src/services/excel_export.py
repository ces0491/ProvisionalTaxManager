"""
Excel export functionality - generates the 11-table tax report
Based on provisional_tax_calc_system.md specifications
"""

import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from decimal import Decimal

from src.config import Config
from src.services.tax_calculator import (
    HOME_OFFICE_CATEGORIES,
    INSURANCE_CATEGORY,
    insurance_deductible_amount,
    DEFAULT_HOME_OFFICE_SQM,
    DEFAULT_HOUSE_TOTAL_SQM,
)


def generate_tax_export(db, Transaction, Category, start_date, end_date, filename):
    """
    Generate Excel file with 11 tables for tax practitioner
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Report"

    # Get all transactions for the period
    transactions = Transaction.query.filter(
        Transaction.date >= start_date,
        Transaction.date <= end_date,
        Transaction.is_deleted == False,  # noqa: E712
        Transaction.is_duplicate == False  # noqa: E712
    ).order_by(Transaction.date).all()

    # Organize by month
    months_in_period = get_months_in_period(start_date, end_date)
    trans_by_month = organize_by_month(transactions, months_in_period)

    # Determine if we need to extrapolate (for incomplete months)
    current_month = datetime.now().date().replace(day=1)
    last_month_in_period = months_in_period[-1]
    needs_extrapolation = last_month_in_period >= current_month

    # Sheet title + audit legend
    office_pct = _office_pct()
    ws['A1'] = f'{Config.BUSINESS_NAME} - Detailed Tax Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = (
        'Amount (R) = amount on the bank statement.  Deductible (R) = qualifying '
        'portion (insurance reduced to building/household-contents; car & life '
        'excluded).  Home-office categories (Interest, Municipal, Insurance) are '
        f'further apportioned at {float(office_pct):.2%} - see Annual Summary and '
        'the Provisional Summary sheet.'
    )
    ws['A2'].font = Font(italic=True, size=9)

    # Build the tables
    row = 4

    # TABLE 1: Monthly Income Summary
    row = write_table1_income_summary(ws, row, trans_by_month, months_in_period, needs_extrapolation)
    row += 2

    # TABLES 2-7: Individual Month Expense Details
    for month in months_in_period:
        row = write_month_detail_table(ws, row, month, trans_by_month[month])
        row += 2

    # TABLE 8: Monthly Business Expense Summary
    row = write_table8_business_summary(ws, row, trans_by_month, months_in_period)
    row += 2

    # TABLE 9: Monthly Net Profit Summary
    row = write_table10_net_profit(ws, row, trans_by_month, months_in_period)
    row += 2

    # TABLE 10: Annual Summary for Tax Calculation
    row = write_table11_annual_summary(ws, row, trans_by_month, start_date, end_date)

    # Practitioner-style summary sheet (placed first so the workbook opens on it)
    write_provisional_summary_sheet(wb, transactions, start_date, end_date)

    # Save workbook to the OS temp directory (portable across Windows/Linux)
    output_path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(output_path)
    return output_path


# Medical categories: not deductible expenses. They feed the medical tax credit
# (applied per member in the tax calculation), so they are reported in their own
# section, never under Expenses.
MEDICAL_CATEGORIES = {'Medical Aid', 'Medical Fees'}


def _office_pct():
    return (DEFAULT_HOME_OFFICE_SQM / DEFAULT_HOUSE_TOTAL_SQM) if DEFAULT_HOUSE_TOTAL_SQM else Decimal('0')


def qualifying_deductible(t):
    """Qualifying deductible portion of a transaction, BEFORE home-office
    apportionment. Uses the signed expense impact (a debit is a positive
    expense, a refund/credit is negative and nets it off). Insurance is reduced
    to its deductible building/household-contents portion; non-business-expense
    transactions (personal, medical, income) return 0. This is the audit bridge
    from the statement amount to what qualifies."""
    if not t.category or t.category.category_type != 'business_expense':
        return Decimal('0')
    amt = -Decimal(str(t.amount))  # debit -> positive expense; refund -> negative
    if t.category.name == INSURANCE_CATEGORY:
        return insurance_deductible_amount(t.description, amt)
    return amt


def claimed_deductible(t, office_pct=None):
    """Final claimed amount: the qualifying portion, with home-office categories
    (interest, rates, insurance) apportioned by the office percentage. Sums to
    the Provisional Summary's total expenses."""
    if office_pct is None:
        office_pct = _office_pct()
    q = qualifying_deductible(t)
    if t.category and t.category.name in HOME_OFFICE_CATEGORIES:
        return q * office_pct
    return q


def write_provisional_summary_sheet(wb, transactions, start_date, end_date):
    """Practitioner-style provisional tax summary sheet.

    Organised into three groups, matching how a tax practitioner reads it:
      - Expenses: deductible business expenses (full amounts), with home-office
        categories rolled into a single apportioned 'Home Office' line.
      - Home Office Apportionment: the building blocks of that line (interest,
        rates, deductible insurance) x office percentage. Insurance is reduced to
        its deductible building/household-contents portion first.
      - Medical: scheme contributions and out-of-pocket fees, shown for
        information only - these support the medical tax credit and are NOT
        deducted from income.
    Personal (non-deductible, non-medical) expenses are excluded entirely.
    """
    ws = wb.create_sheet('Provisional Summary', 0)

    office_sqm = DEFAULT_HOME_OFFICE_SQM
    house_sqm = DEFAULT_HOUSE_TOTAL_SQM
    office_pct = (office_sqm / house_sqm) if house_sqm else Decimal('0')

    bold = Font(bold=True)
    title = Font(bold=True, size=14)
    italic = Font(italic=True, size=9)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    def money(row, value):
        cell = ws.cell(row, 2, float(value))
        cell.number_format = '#,##0.00'
        return cell

    # --- Aggregate ---
    income_by_month = {}      # 'YYYY-MM' -> Decimal
    expense_by_category = {}  # deductible (non-home-office, non-medical) business expense -> Decimal
    home_office_base = {}     # home-office category -> deductible base Decimal
    medical_by_category = {}  # medical category -> Decimal (informational, not deducted)

    for t in transactions:
        if t.is_deleted or t.is_duplicate:
            continue
        cat = t.category.name if t.category else 'Uncategorized'
        ctype = t.category.category_type if t.category else 'personal_expense'
        amt = abs(Decimal(str(t.amount)))

        # Medical is pulled out regardless of stored type (credit, not a deduction)
        if cat in MEDICAL_CATEGORIES:
            medical_by_category[cat] = medical_by_category.get(cat, Decimal('0')) + amt
        elif ctype == 'income':
            key = t.date.strftime('%Y-%m')
            income_by_month[key] = income_by_month.get(key, Decimal('0')) + amt
        elif ctype == 'business_expense':
            # Signed, insurance-split deductible (refunds net off); same helper
            # the detailed Tax Report uses, so the two sheets reconcile.
            q = qualifying_deductible(t)
            if cat in HOME_OFFICE_CATEGORIES:
                home_office_base[cat] = home_office_base.get(cat, Decimal('0')) + q
            else:
                expense_by_category[cat] = expense_by_category.get(cat, Decimal('0')) + q
        # personal / excluded (non-medical) expenses are intentionally omitted

    ho_subtotal = sum(home_office_base.values(), Decimal('0'))
    home_office_deduction = (ho_subtotal * office_pct).quantize(Decimal('0.01'))
    total_income = sum(income_by_month.values(), Decimal('0'))
    total_expenses = sum(expense_by_category.values(), Decimal('0')) + home_office_deduction
    net_profit = total_income - total_expenses

    # --- Write ---
    r = 1
    ws.cell(r, 1, f'{Config.BUSINESS_NAME} - Provisional Tax Summary').font = title
    r += 1
    ws.cell(r, 1, f'Tax Reference: {Config.TAX_REFERENCE}')
    r += 1
    ws.cell(r, 1, f"Period: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}")
    r += 2

    # Income
    c = ws.cell(r, 1, 'Income (by month)')
    c.font = bold
    c.fill = header_fill
    r += 1
    for key in sorted(income_by_month):
        ws.cell(r, 1, datetime.strptime(key, '%Y-%m').strftime('%b %Y'))
        money(r, income_by_month[key])
        r += 1
    ws.cell(r, 1, 'Total Income').font = bold
    money(r, total_income).font = bold
    r += 2

    # Expenses (deductible)
    c = ws.cell(r, 1, 'Expenses (deductible)')
    c.font = bold
    c.fill = header_fill
    r += 1
    for cat in sorted(expense_by_category):
        ws.cell(r, 1, cat)
        money(r, expense_by_category[cat])
        r += 1
    ws.cell(r, 1, 'Home Office (apportioned)')
    money(r, home_office_deduction)
    r += 1
    ws.cell(r, 1, 'Total Expenses').font = bold
    money(r, total_expenses).font = bold
    r += 2

    ws.cell(r, 1, 'Net Profit (period)').font = bold
    money(r, net_profit).font = bold
    r += 2

    # Home Office Apportionment
    c = ws.cell(r, 1, 'Home Office Apportionment')
    c.font = bold
    c.fill = header_fill
    r += 1
    for cat in sorted(home_office_base):
        label = 'Insurance (building/contents)' if cat == INSURANCE_CATEGORY else cat
        ws.cell(r, 1, label)
        money(r, home_office_base[cat])
        r += 1
    ws.cell(r, 1, 'Subtotal').font = bold
    money(r, ho_subtotal).font = bold
    r += 1
    ws.cell(r, 1, 'Office size (m²)')
    ws.cell(r, 2, float(office_sqm))
    r += 1
    ws.cell(r, 1, 'Home size (m²)')
    ws.cell(r, 2, float(house_sqm))
    r += 1
    ws.cell(r, 1, 'Percentage claim')
    pct_cell = ws.cell(r, 2, float(round(office_pct, 4)))
    pct_cell.number_format = '0.00%'
    r += 1
    ws.cell(r, 1, 'Home Office deduction').font = bold
    money(r, home_office_deduction).font = bold
    r += 2

    # Medical (tax credit - not a deduction)
    c = ws.cell(r, 1, 'Medical (tax credit - not a deduction)')
    c.font = bold
    c.fill = header_fill
    r += 1
    for cat in sorted(medical_by_category):
        ws.cell(r, 1, cat)
        money(r, medical_by_category[cat])
        r += 1
    ws.cell(r, 1,
            'Not deducted from income. Supports the medical tax credit, '
            'applied per member in the tax calculation.').font = italic
    r += 1

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 16

    # Open the workbook on this summary sheet, not the detailed Tax Report.
    wb.active = wb.index(ws)
    return ws


def get_months_in_period(start_date, end_date):
    """Get list of month start dates in the period"""
    months = []
    current = start_date.replace(day=1)
    end = end_date.replace(day=1)

    while current <= end:
        months.append(current)
        # Move to next month
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    return months


def organize_by_month(transactions, months):
    """Organize transactions by month"""
    trans_by_month = {month: [] for month in months}

    for trans in transactions:
        month_start = trans.date.replace(day=1)
        if month_start in trans_by_month:
            trans_by_month[month_start].append(trans)

    return trans_by_month


def write_table1_income_summary(ws, start_row, trans_by_month, months, needs_extrapolation):
    """Table 1: Monthly Income Summary"""
    row = start_row

    # Header
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'Table 1: Monthly Income Summary'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    # Column headers
    headers = ['Month', 'Description', 'Amount (R)', 'Source']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    row += 1

    total_income = Decimal('0')

    for month in months:
        # Get income transactions for this month
        income_trans = [t for t in trans_by_month[month]
                       if t.category and t.category.category_type == 'income']

        if income_trans:
            month_name = month.strftime('%b-%y')
            for trans in income_trans:
                ws.cell(row=row, column=1).value = month_name
                ws.cell(row=row, column=2).value = trans.description
                ws.cell(row=row, column=3).value = float(trans.amount)
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                ws.cell(row=row, column=4).value = f"Statement {trans.statement.account.name}"
                total_income += trans.amount
                row += 1

    # Total row
    ws.cell(row=row, column=1).value = 'TOTAL'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3).value = float(total_income)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    row += 1

    return row


def write_month_detail_table(ws, start_row, month, transactions):
    """Tables 2-7: Individual Month Expense Details"""
    row = start_row
    month_name = month.strftime('%B %Y')

    # Header
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f'Business Expenses for {month_name}'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    # Column headers: statement amount AND the qualifying deductible portion
    headers = ['Category', 'Description', 'Amount (R)', 'Deductible (R)', 'Date', 'Source']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    row += 1

    business_trans = [t for t in transactions
                     if t.category and t.category.category_type == 'business_expense']
    gross_total = Decimal('0')
    deductible_total = Decimal('0')

    for trans in business_trans:
        gross = -Decimal(str(trans.amount))  # signed expense impact (refund = negative)
        ded = qualifying_deductible(trans)
        ws.cell(row=row, column=1).value = trans.category.name
        ws.cell(row=row, column=2).value = trans.description
        ws.cell(row=row, column=3).value = float(gross)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4).value = float(ded)
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).value = trans.date.strftime('%d-%b')
        ws.cell(row=row, column=6).value = f"{trans.statement.account.account_type}"
        gross_total += gross
        deductible_total += ded
        row += 1

    # Subtotal (statement amount and qualifying deductible)
    ws.cell(row=row, column=1).value = 'SUBTOTAL'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3).value = float(gross_total)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4).value = float(deductible_total)
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=4).font = Font(bold=True)
    row += 1

    # Personal/non-deductible expenses are intentionally not listed.
    return row


def write_table8_business_summary(ws, start_row, trans_by_month, months):
    """Table 8: Monthly Business Expense Summary"""
    row = start_row

    # Header
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = 'Table 8: Deductible Business Expenses by Category and Month'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1).value = (
        'Qualifying amounts (insurance split applied), before home-office apportionment.'
    )
    ws.cell(row=row, column=1).font = Font(italic=True, size=9)
    row += 1

    # Column headers
    ws.cell(row=row, column=1).value = 'Category'
    ws.cell(row=row, column=1).font = Font(bold=True)

    for col, month in enumerate(months, start=2):
        ws.cell(row=row, column=col).value = month.strftime('%b-%y')
        ws.cell(row=row, column=col).font = Font(bold=True)

    ws.cell(row=row, column=len(months) + 2).value = 'Total'
    ws.cell(row=row, column=len(months) + 2).font = Font(bold=True)
    row += 1

    # Category rows (only categories that actually have deductible amounts)
    from src.services.categorizer import CATEGORIES
    business_categories = [cat['name'] for cat in CATEGORIES.values() if cat['type'] == 'business_expense']

    for cat_name in business_categories:
        cat_cells = []
        row_total = Decimal('0')
        for month in months:
            month_total = sum(
                (qualifying_deductible(t) for t in trans_by_month[month]
                 if t.category and t.category.name == cat_name),
                Decimal('0'),
            )
            cat_cells.append(month_total)
            row_total += month_total
        if row_total == 0:
            continue  # skip categories with no activity in the period
        ws.cell(row=row, column=1).value = cat_name
        for col, month_total in enumerate(cat_cells, start=2):
            ws.cell(row=row, column=col).value = float(month_total)
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=len(months) + 2).value = float(row_total)
        ws.cell(row=row, column=len(months) + 2).number_format = '#,##0.00'
        row += 1

    # Monthly totals row
    ws.cell(row=row, column=1).value = 'MONTHLY TOTALS'
    ws.cell(row=row, column=1).font = Font(bold=True)
    grand_total = Decimal('0')
    for col, month in enumerate(months, start=2):
        month_total = sum(
            (qualifying_deductible(t) for t in trans_by_month[month]
             if t.category and t.category.category_type == 'business_expense'),
            Decimal('0'),
        )
        ws.cell(row=row, column=col).value = float(month_total)
        ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=col).font = Font(bold=True)
        grand_total += month_total
    ws.cell(row=row, column=len(months) + 2).value = float(grand_total)
    ws.cell(row=row, column=len(months) + 2).number_format = '#,##0.00'
    ws.cell(row=row, column=len(months) + 2).font = Font(bold=True)

    row += 1
    return row


def write_table10_net_profit(ws, start_row, trans_by_month, months):
    """Table 9: Monthly Net Profit Summary (deductible expenses, after home-office apportionment)."""
    row = start_row
    office_pct = _office_pct()

    # Header
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'Table 9: Monthly Net Profit Summary'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    # Column headers
    headers = ['Month', 'Income (R)', 'Deductible Expenses (R)', 'Net Profit (R)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    row += 1

    period_income = Decimal('0')
    period_expenses = Decimal('0')

    for month in months:
        # Income
        income_trans = [t for t in trans_by_month[month]
                       if t.category and t.category.category_type == 'income']
        month_income = sum([t.amount for t in income_trans], Decimal('0'))

        # Deductible expenses (insurance split + home-office apportionment)
        month_expenses = sum(
            (claimed_deductible(t, office_pct) for t in trans_by_month[month]
             if t.category and t.category.category_type == 'business_expense'),
            Decimal('0'),
        )

        net_profit = month_income - month_expenses

        ws.cell(row=row, column=1).value = month.strftime('%b-%y')
        ws.cell(row=row, column=2).value = float(month_income)
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3).value = float(month_expenses)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4).value = float(net_profit)
        ws.cell(row=row, column=4).number_format = '#,##0.00'

        period_income += month_income
        period_expenses += month_expenses
        row += 1

    # Total row
    ws.cell(row=row, column=1).value = 'TOTAL'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = float(period_income)
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).value = float(period_expenses)
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4).value = float(period_income - period_expenses)
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=4).font = Font(bold=True)
    row += 1

    return row


def write_table11_annual_summary(ws, start_row, trans_by_month, start_date, end_date):
    """Table 11: Annual Summary for Tax Calculation"""
    row = start_row

    # Header
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = 'Table 10: Annual Summary for Tax Calculation'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    office_pct = _office_pct()

    # Calculate totals from all months, bridging gross -> qualifying -> claimed.
    all_income = Decimal('0')
    qualifying = Decimal('0')        # deductible portion before apportionment
    ho_qualifying = Decimal('0')     # the home-office part of qualifying

    for month_trans in trans_by_month.values():
        all_income += sum(
            (t.amount for t in month_trans
             if t.category and t.category.category_type == 'income'),
            Decimal('0'),
        )
        for t in month_trans:
            if t.category and t.category.category_type == 'business_expense':
                q = qualifying_deductible(t)
                qualifying += q
                if t.category.name in HOME_OFFICE_CATEGORIES:
                    ho_qualifying += q

    # Home-office apportionment removes the non-office share of home-office items.
    apportionment_adj = (ho_qualifying * (Decimal('1') - office_pct)).quantize(Decimal('0.01'))
    deductible = (qualifying - apportionment_adj).quantize(Decimal('0.01'))
    net_profit = (all_income - deductible).quantize(Decimal('0.01'))

    ws.cell(row=row, column=1).value = 'Description'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = 'Amount (R)'
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1

    period = f'{start_date.strftime("%b %Y")} - {end_date.strftime("%b %Y")}'

    def line(label, value, bold=False):
        nonlocal row
        ws.cell(row=row, column=1).value = label
        c = ws.cell(row=row, column=2)
        c.value = float(value)
        c.number_format = '#,##0.00'
        if bold:
            ws.cell(row=row, column=1).font = Font(bold=True)
            c.font = Font(bold=True)
        row += 1

    line(f'Period Income ({period})', all_income)
    line('Qualifying expenses (before home-office apportionment)', qualifying)
    line(f'Less home-office apportionment ({float(1 - office_pct):.2%} of home-office items)', -apportionment_adj)
    line('Deductible expenses (claimed)', deductible, bold=True)
    line('Period Net Profit', net_profit, bold=True)
    row += 1

    line('Annualized Income (projected x2)', all_income * 2, bold=True)
    line('Annualized Deductible Expenses (x2)', deductible * 2, bold=True)
    line('Annualized Net Profit (for provisional tax)', net_profit * 2, bold=True)

    return row
